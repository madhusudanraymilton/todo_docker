import io
import base64
from datetime import datetime, timedelta

from odoo import models, fields, api
from odoo.exceptions import UserError


class ShiftAssignment(models.Model):
    _name = 'shift.assignment'
    _description = 'Shift Assignment'

    employee_id = fields.Many2one('hr.employee', required=True, ondelete='cascade', index=True)
    date = fields.Date(required=True, index=True)
    shift_id = fields.Many2one('resource.calendar', string='Shift', ondelete='set null')
    shift_code = fields.Char(related='shift_id.shift_code', store=True, readonly=True)

    _sql_constraints = [
        ('unique_employee_date', 'UNIQUE(employee_id, date)',
         'Only one shift per employee per day.'),
    ]

    # ── Sync helper ───────────────────────────────────────────────────────────

    def _sync_employee_work_schedule(self, employee_id, date_str, shift_id):
        """
        If *date_str* equals today, write the shift's resource.calendar onto
        hr.employee.resource_calendar_id so the employee Work Schedule always
        reflects today's rostered shift.

        Clearing a cell (shift_id=False) is intentionally ignored — there is
        no meaningful schedule to revert to when a shift is removed.
        """
        if not shift_id:
            return

        today_str = fields.Date.today().strftime('%Y-%m-%d')
        if str(date_str) != today_str:
            return

        employee = self.env['hr.employee'].browse(employee_id)
        if not employee.exists():
            return

        # resource.calendar IS the work schedule model; shift_id is its id.
        if employee.resource_calendar_id.id != shift_id:
            employee.sudo().write({'resource_calendar_id': shift_id})

    # ── Dashboard data helpers ─────────────────────────────────────────────────

    @api.model
    def get_roster_data(self, date_from, date_to):
        """Return {emp_id: {date_str: {shift_id, shift_code}}}"""
        recs = self.search_read(
            [('date', '>=', date_from), ('date', '<=', date_to)],
            ['employee_id', 'date', 'shift_id', 'shift_code'],
        )
        result = {}
        for r in recs:
            eid = r['employee_id'][0]
            ds = str(r['date'])
            result.setdefault(eid, {})[ds] = {
                'shift_id': r['shift_id'][0] if r['shift_id'] else False,
                'shift_code': r['shift_code'] or '',
            }
        return result

    @api.model
    def set_shift(self, employee_id, date, shift_id):
        """
        Create / update / delete a single cell assignment (manual dashboard pick).

        After saving the roster record, syncs hr.employee.resource_calendar_id
        when the assigned date is today — so the employee Work Schedule always
        matches today's rostered shift in real time.
        """
        rec = self.search([('employee_id', '=', employee_id), ('date', '=', date)], limit=1)
        if rec:
            if shift_id:
                rec.write({'shift_id': shift_id})
            else:
                rec.unlink()
        elif shift_id:
            self.create({'employee_id': employee_id, 'date': date, 'shift_id': shift_id})

        # Auto-sync Work Schedule only when this cell is for today
        self._sync_employee_work_schedule(employee_id, date, shift_id)

        return True

    @api.model
    def set_shift_bulk(self, assignments, overwrite=False):
        """
        Bulk-set shift assignments.
        assignments: list of {employee_id, date, shift_id}
        overwrite=False: only fills blank cells (default shift behaviour)
        overwrite=True:  upserts all cells (explicit multi-select apply)
        Returns {saved: N}
        """
        if not assignments:
            return {'saved': 0}

        today_str = fields.Date.today().strftime('%Y-%m-%d')
        emp_ids   = list({a['employee_id'] for a in assignments})
        dates     = list({a['date'] for a in assignments})

        # Fetch existing records for these (emp, date) pairs
        existing     = self.search([('employee_id', 'in', emp_ids), ('date', 'in', dates)])
        existing_map = {(r.employee_id.id, str(r.date)): r for r in existing}

        to_create    = []
        today_entries = []
        saved = 0

        for a in assignments:
            key = (a['employee_id'], a['date'])
            rec = existing_map.get(key)

            if rec:
                if overwrite:
                    rec.write({'shift_id': a['shift_id']})
                    saved += 1
                    if a['date'] == today_str:
                        today_entries.append(a)
                # else: skip existing (default shift mode)
            else:
                to_create.append({
                    'employee_id': a['employee_id'],
                    'date':        a['date'],
                    'shift_id':    a['shift_id'],
                })
                if a['date'] == today_str:
                    today_entries.append(a)

        if to_create:
            self.create(to_create)
            saved += len(to_create)

        for a in today_entries:
            self._sync_employee_work_schedule(a['employee_id'], a['date'], a['shift_id'])

        return {'saved': saved}

    # ── XLSX Export ────────────────────────────────────────────────────────────

    @api.model
    def export_xlsx(self, date_from, date_to):
        """Build roster XLSX and return base64 string."""
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
            from openpyxl.utils import get_column_letter
        except ImportError:
            raise UserError('openpyxl is not installed. Run: pip install openpyxl')

        d_from = datetime.strptime(date_from, '%Y-%m-%d').date()
        d_to = datetime.strptime(date_to, '%Y-%m-%d').date()
        dates = []
        cur = d_from
        while cur <= d_to:
            dates.append(cur)
            cur += timedelta(days=1)

        employees = self.env['hr.employee'].search_read(
            [], ['id', 'name', 'barcode', 'department_id', 'job_id'], order='name'
        )
        roster = self.get_roster_data(date_from, date_to)

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'Shift Roster'

        # Styles
        thin = Side(style='thin', color='D1D5DB')
        border = Border(left=thin, right=thin, top=thin, bottom=thin)
        hdr_font = Font(name='Arial', bold=True, color='FFFFFF', size=10)
        emp_fill = PatternFill('solid', start_color='4C1D95')
        date_fill = PatternFill('solid', start_color='7C3AED')
        alt_fill = PatternFill('solid', start_color='F5F3FF')
        data_font = Font(name='Arial', size=9)
        center = Alignment(horizontal='center', vertical='center', wrap_text=False)
        left_al = Alignment(horizontal='left', vertical='center')

        all_headers = ['Employee Name', 'Employee ID'] + [d.strftime('%d/%m/%Y') for d in dates]

        for ci, h in enumerate(all_headers, 1):
            c = ws.cell(row=1, column=ci, value=h)
            c.font = hdr_font
            c.fill = emp_fill if ci <= 2 else date_fill
            c.alignment = center
            c.border = border
        ws.row_dimensions[1].height = 22

        for ri, emp in enumerate(employees, 2):
            is_alt = ri % 2 == 0
            for ci, val in enumerate([emp['name'], emp['barcode'] or ''], 1):
                c = ws.cell(row=ri, column=ci, value=val)
                c.font = data_font
                c.alignment = left_al
                c.border = border
                if is_alt:
                    c.fill = alt_fill

            emp_assignments = roster.get(emp['id'], {})
            for di, date_obj in enumerate(dates):
                ci = 3 + di
                sc = emp_assignments.get(date_obj.strftime('%Y-%m-%d'), {}).get('shift_code', '')
                c = ws.cell(row=ri, column=ci, value=sc)
                c.font = data_font
                c.alignment = center
                c.border = border
                if is_alt:
                    c.fill = alt_fill

        ws.column_dimensions['A'].width = 30
        ws.column_dimensions['B'].width = 16
        for i in range(len(dates)):
            ws.column_dimensions[get_column_letter(3 + i)].width = 10

        ws.freeze_panes = 'C2'

        buf = io.BytesIO()
        wb.save(buf)
        return base64.b64encode(buf.getvalue()).decode()

    # ── XLSX Import ────────────────────────────────────────────────────────────

    @api.model
    def import_xlsx(self, file_b64):
        """
        Parse an uploaded XLSX (same format as export) and upsert assignments.
        Columns: Employee Name | Employee ID | date1 | date2 ...

        For every cell whose date equals today, hr.employee.resource_calendar_id
        is updated to the rostered shift — same rule as a manual single-cell
        pick on the dashboard.

        Returns {'imported': N, 'skipped': N, 'errors': [...]}
        """
        try:
            import openpyxl
        except ImportError:
            raise UserError('openpyxl is not installed. Run: pip install openpyxl')

        raw = base64.b64decode(file_b64)
        wb = openpyxl.load_workbook(io.BytesIO(raw), data_only=True)
        ws = wb.active

        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            raise UserError('The uploaded file is empty.')

        header_row = rows[0]
        date_cols = []
        for ci, h in enumerate(header_row):
            if ci < 2:
                continue
            if h is None:
                continue
            try:
                if isinstance(h, str):
                    date_obj = datetime.strptime(h.strip(), '%d/%m/%Y').date()
                elif hasattr(h, 'date'):
                    date_obj = h.date() if hasattr(h, 'date') and callable(h.date) else h
                else:
                    date_obj = h
                date_cols.append((ci, date_obj.strftime('%Y-%m-%d')))
            except Exception:
                continue

        if not date_cols:
            raise UserError('No valid date columns found. Expected format: dd/mm/yyyy')

        all_shifts = self.env['resource.calendar'].search_read([], ['id', 'shift_code'])
        shift_map = {s['shift_code'].strip().upper(): s['id'] for s in all_shifts if s['shift_code']}

        all_employees = self.env['hr.employee'].search_read([], ['id', 'name', 'barcode'])
        emp_by_name = {e['name'].strip().lower(): e['id'] for e in all_employees}
        emp_by_barcode = {str(e['barcode']).strip(): e['id'] for e in all_employees if e['barcode']}

        today_str = fields.Date.today().strftime('%Y-%m-%d')

        imported = 0
        skipped = 0
        errors = []

        for row_i, row in enumerate(rows[1:], start=2):
            emp_name = row[0]
            emp_barcode = row[1]

            emp_id = None
            if emp_barcode and str(emp_barcode).strip() in emp_by_barcode:
                emp_id = emp_by_barcode[str(emp_barcode).strip()]
            elif emp_name and str(emp_name).strip().lower() in emp_by_name:
                emp_id = emp_by_name[str(emp_name).strip().lower()]

            if not emp_id:
                errors.append(f'Row {row_i}: Employee not found: {emp_name} / {emp_barcode}')
                skipped += 1
                continue

            for ci, date_str in date_cols:
                cell_val = row[ci] if ci < len(row) else None
                if cell_val is None or str(cell_val).strip() == '':
                    # Clear the assignment — no Work Schedule sync on clear
                    rec = self.search(
                        [('employee_id', '=', emp_id), ('date', '=', date_str)], limit=1
                    )
                    if rec:
                        rec.unlink()
                    continue

                code = str(cell_val).strip().upper()
                shift_id = shift_map.get(code)
                if not shift_id:
                    errors.append(f'Row {row_i}, date {date_str}: Unknown shift code "{cell_val}"')
                    skipped += 1
                    continue

                # Upsert the shift.assignment record
                rec = self.search(
                    [('employee_id', '=', emp_id), ('date', '=', date_str)], limit=1
                )
                if rec:
                    rec.write({'shift_id': shift_id})
                else:
                    self.create({'employee_id': emp_id, 'date': date_str, 'shift_id': shift_id})

                # Sync Work Schedule only when this cell is for today
                if date_str == today_str:
                    self._sync_employee_work_schedule(emp_id, date_str, shift_id)

                imported += 1

        return {'imported': imported, 'skipped': skipped, 'errors': errors[:20]}
